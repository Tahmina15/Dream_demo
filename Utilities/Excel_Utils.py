import openpyxl


# Method for how many rows are using for testing
def get_rows_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row


# Method for how many columns are using for testing
def get_column_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_column


# Method for read the data from Excel sheet
def read_data(file, sheet_name, r, c):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row=r, column=c).value


# Write something in Excel throughout testing
def write_data(file, sheet_name, r, c, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row=r, column=c).value = data
    workbook.save(file)


